import os
import io
import re
import json
import sqlite3
import logging
import random
import traceback
import collections
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from functools import wraps

import pandas as pd
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from dateutil import parser

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, send_file, jsonify, Response, session
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from sqlalchemy import func, case, text
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature


# Configuração de logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__) 



app.config['SECRET_KEY'] = 'your_secret_key'  # Substitua por uma chave segura em produção
#app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///E:/Python/Abertura_de_Atendimento/instance/tickets.db?timeout=10'  # Aumentado o timeout
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['UPLOAD_FOLDER'] = 'E:/Python/Abertura_de_Atendimento/instance/uploads'  # Usando caminho relativo

# Criar diretório de uploads se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Configuração do Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Inicialização do serializer para tokens
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# =======================
# Permissões por URL
# =======================
class Permission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    AUX1 = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    url = db.Column(db.String(200), nullable=False)
    user = db.relationship('User', backref='permissions')



def _path_allowed(allowed_urls, path):
    """
    Suporta correspondência exata e wildcard com '*'.
    Converte o padrão de URL em uma expressão regular para uma correspondência robusta.
    """
    # Normaliza o caminho solicitado, removendo a barra final se houver
    path = path.rstrip('/') or '/'

    for url_pattern in allowed_urls:
        # Normaliza o padrão da permissão
        pattern = (url_pattern or '/').rstrip('/') or '/'
        
        # Converte o padrão com wildcard em uma expressão regular
        # Ex: /ticket/*/pause -> ^/ticket/[^/]+/pause$
        regex_pattern = re.escape(pattern).replace('\\*', '[^/]+')
        
        # Garante que a correspondência seja exata do início ao fim
        if re.fullmatch(f"^{regex_pattern}$", path):
            return True
            
    return False

def permission_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Você precisa estar logado para acessar esta página.', 'error')
            return redirect(url_for('login'))

        if getattr(current_user, 'is_admin', False):
            logger.debug("Usuário é admin, acesso total permitido.")
            return f(*args, **kwargs)

        urls_basicas = ['/', '/logout', '/sem_permissao']
        path_solicitado = request.path.rstrip('/') or '/'

        if path_solicitado in urls_basicas:
            logger.debug(f"Acesso permitido à URL básica: {path_solicitado}")
            return f(*args, **kwargs)

        if not current_user.permissions:
            flash('Acesso negado: nenhuma permissão configurada para este usuário.', 'error')
            logger.debug("Nenhuma permissão configurada, redirecionando para sem_permissao")
            return redirect(url_for('sem_permissao'))

        allowed_urls = [p.url for p in current_user.permissions]
        logger.debug(f"URLs permitidas: {allowed_urls}, Caminho solicitado: {request.path}")
        
        if not _path_allowed(allowed_urls, request.path):
            flash('Acesso negado: você não tem permissão para esta página.', 'error')
            logger.debug("Permissão negada, redirecionando para sem_permissao")
            return redirect(url_for('sem_permissao'))
            
        return f(*args, **kwargs)
    return decorated_function

# =======================================
# FUNÇÃO AUXILIAR PARA USAR NOS TEMPLATES (LUGAR CORRETO)
# =======================================
@app.context_processor
def utility_processor():
    def has_permission(url_path):
        if getattr(current_user, 'is_admin', False):
            return True
        
        if not current_user.is_authenticated:
            return False
            
        allowed_urls = [p.url for p in current_user.permissions]
        return _path_allowed(allowed_urls, url_path)

    return dict(has_permission=has_permission)


# =======================
# Modelos da aplicação
# =======================
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Ticket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    os = db.Column(db.String(50), nullable=True)
    csr = db.Column(db.String(50), nullable=True)
    username = db.Column(db.String(50), nullable=False)
    tool = db.Column(db.String(50), nullable=False)
    subject = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    priority = db.Column(db.String(20), nullable=False, default='Baixo')
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(ZoneInfo("America/Sao_Paulo")))
    status = db.Column(db.String(20), nullable=False, default='open')
    closed_at = db.Column(db.DateTime)
    
    status_sla = db.Column(db.String(20), nullable=False, default='running') # Estados: 'running', 'paused'
    paused_at = db.Column(db.DateTime, nullable=True) # Momento em que a última pausa começou
    total_paused_duration = db.Column(db.Integer, default=0) # Duração total da pausa em segundos
    
    files = db.relationship('File', backref='ticket', lazy=True)
    comments = db.relationship('Comment', backref='ticket', lazy=True)
    TR = db.Column(db.String(50))
    CR = db.Column(db.String(50))
    AUX1 = db.Column(db.String(50))
    AUX2 = db.Column(db.String(50))
    AUX3 = db.Column(db.String(50))
    


class File(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    filename = db.Column(db.String(100), nullable=False)

class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    text = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=lambda: datetime.now(ZoneInfo("America/Sao_Paulo")))
    files = db.relationship('CommentFile', backref='comment', lazy=True)

class CommentFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    comment_id = db.Column(db.Integer, db.ForeignKey('comment.id'), nullable=False)
    filename = db.Column(db.String(100), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def get_csr_db_connection():
    # Define o caminho absoluto para o arquivo do banco de dados
    db_path = r'E:\Python\Abertura_de_Atendimento\instance\csr.db'
    
    # Conecta usando o caminho absoluto
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def save_comment_to_excel(url, username, comment):
    file_path = 'E:/Python/Coleta-Site/exportado_formatado.xlsx'
    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet.max_row == 1 and sheet['A1'].value is None:
            sheet.append(['URL', 'Username', 'Comment', 'Timestamp'])
        sheet.append([url, username, comment, datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat()])
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar comentário no Excel: {e}")
        return False

# =======================
# Autenticação
# =======================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Login bem-sucedido!', 'success')
            logger.debug(f"Usuário {username} logado com sucesso")
            return redirect(url_for('index'))
        else:
            flash('Usuário ou senha inválidos.', 'error')
            logger.debug(f"Falha no login para usuário {username}")
    return render_template('login.html')

@app.route('/logout')
@login_required
@permission_required
def logout():
    logout_user()
    flash('Você foi desconectado.', 'success')
    return redirect(url_for('login'))

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password_request():
    if request.method == 'POST':
        username = request.form['username']
        user = User.query.filter_by(username=username).first()
        if user:
            token = serializer.dumps(user.username, salt='reset-password')
            flash(f'Um link para redefinir sua senha foi enviado (simulação). Use o token: {token}', 'info')
            return redirect(url_for('reset_password', token=token))
        else:
            flash('Usuário não encontrado.', 'error')
    return render_template('reset_password_request.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        username = serializer.loads(token, salt='reset-password', max_age=3600)
        user = User.query.filter_by(username=username).first()
        if not user:
            flash('Usuário inválido ou token expirado.', 'error')
            return redirect(url_for('login'))
        if request.method == 'POST':
            new_password = request.form['new_password']
            user.set_password(new_password)
            db.session.commit()
            flash('Senha redefinida com sucesso. Faça login com a nova senha.', 'success')
            return redirect(url_for('login'))
        return render_template('reset_password.html', token=token)
    except (SignatureExpired, BadSignature):
        flash('O token está expirado ou é inválido.', 'error')
        return redirect(url_for('login'))

# =======================
# Rota para Acesso Negado
# =======================
@app.route('/sem_permissao')
def sem_permissao():
    return render_template('sem_permissao.html'), 403

# =======================
# Admin (com gestão de permissões)
# =======================
@app.route('/admin', methods=['GET', 'POST'])
@login_required
@permission_required
def admin():
    if not current_user.is_admin:
        flash('Acesso negado: apenas administradores podem acessar esta página.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            username = request.form['username']
            password = request.form['password']
            is_admin = 'is_admin' in request.form
            if User.query.filter_by(username=username).first():
                flash('Usuário já existe.', 'error')
            else:
                new_user = User(username=username, is_admin=is_admin)
                new_user.set_password(password)
                db.session.add(new_user)
                db.session.commit()
                flash('Usuário criado com sucesso!', 'success')

        elif action == 'delete':
            user_id = request.form['user_id']
            user = db.session.get(User, int(user_id))
            if user.is_admin and User.query.filter_by(is_admin=True).count() == 1:
                flash('Não é possível excluir o último administrador.', 'error')
            else:
                db.session.delete(user)
                db.session.commit()
                flash('Usuário excluído com sucesso!', 'success')

        elif action == 'add_permission':
            user_id = int(request.form['user_id'])
            url = request.form['url'].strip()
            if not Permission.query.filter_by(AUX1=user_id, url=url).first():
                db.session.add(Permission(AUX1=user_id, url=url))
                db.session.commit()
                flash('Permissão adicionada com sucesso!', 'success')
            else:
                flash('Esta permissão já existe para o usuário.', 'warning')

        elif action == 'remove_permission':
            perm_id = int(request.form['perm_id'])
            perm = db.session.get(Permission, perm_id)
            if perm:
                db.session.delete(perm)
                db.session.commit()
                flash('Permissão removida com sucesso!', 'success')
            else:
                flash('Permissão não encontrada.', 'error')
        
        return redirect(url_for('admin')) # Adicionado para recarregar a página após uma ação POST

    users = User.query.all()
    permissions = Permission.query.all()
    return render_template('admin.html', users=users, permissions=permissions)

# =======================
# Rotas da aplicação
# =======================


@app.template_filter('format_datetime')
def format_datetime_filter(value):
    if not value:
        return ""
    try:
        # Tenta converter de string para objeto datetime, se necessário
        if isinstance(value, str):
            dt_object = parser.parse(value)
        else:
            dt_object = value  # Assume que já é um objeto datetime
            
        # Garante que a data está em America/Sao_Paulo
        if dt_object.tzinfo is None:
            dt_object = dt_object.replace(tzinfo=ZoneInfo("America/Sao_Paulo"))
        else:
            dt_object = dt_object.astimezone(ZoneInfo("America/Sao_Paulo"))
            
        # Formata o objeto datetime para o formato desejado
        return dt_object.strftime('%d/%m/%Y %H:%M')
    except (ValueError, TypeError):
        logger.warning(f"Não foi possível formatar o valor de data: {value}")
        return str(value)  # Retorna o valor original como string se a formatação falhar

# Configuração do Flask-Login


@app.route('/csr')
@login_required
def lista_csr2():
    conn = get_csr_db_connection()
    cursor = conn.cursor()
    # Depuração: Selecionar todos os tickets para verificar as datas
    cursor.execute('SELECT * FROM tickets')
    all_rows = cursor.fetchall()
    logging.debug(f"Todos os tickets: {all_rows}")
    
    # Selecionar o mais recente
    cursor.execute('SELECT * FROM tickets ORDER BY "Created Time" DESC LIMIT 500')
    rows = cursor.fetchall()
    logging.debug(f"Ticket mais recente: {rows}")
    conn.close()
    
    if not rows:
        return render_template('lista_csr2.html', rows=[], error="Nenhum ticket encontrado.")
    
    return render_template('lista_csr2.html', rows=rows)

@app.errorhandler(404)
def page_not_found(error):
    return render_template('404.html'), 404





@app.route('/lista_csr')
@login_required
@permission_required
def lista_csr():
    status_filter = request.args.get('status', default=None)
    logger.debug(f"Filtro de status recebido: {status_filter}")
    
    try:
        # Use o modelo Ticket do SQLAlchemy diretamente
        query = Ticket.query
        
        if status_filter:
            query = query.filter(Ticket.status == status_filter)
            
        # Filtra para mostrar apenas tickets que têm um valor em 'csr'
        # Isso garante que a página mostre apenas o que se espera dela.
        tickets_com_csr = query.filter(Ticket.csr != None, Ticket.csr != '').all()
        
        logger.debug(f"Tickets com CSR retornados: {len(tickets_com_csr)}")
        
        # O template 'csr.html' provavelmente espera uma lista de dicionários,
        # então vamos converter os objetos Ticket.
        tickets_dict = [
            {
                'id': t.id, 
                'os': t.os, 
                'csr': t.csr, 
                'username': t.username, 
                'tool': t.tool, 
                'subject': t.subject,
                'description': t.description,
                'priority': t.priority,
                'created_at': t.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'status': t.status,
                'closed_at': t.closed_at.strftime('%Y-%m-%d %H:%M:%S') if t.closed_at else 'N/A'
            } for t in tickets_com_csr
        ]
        
        return render_template('csr.html', rows=tickets_dict)

    except Exception as e:
        logger.error(f"Erro ao buscar lista de CSRs: {e}\n{traceback.format_exc()}")
        flash('Ocorreu um erro ao carregar a lista de CSRs.', 'error')
        return render_template('csr.html', rows=[])

############################################      ROTA OK   ###########################################################################################################

@app.route('/relatorio_ferramentas')
@login_required
@permission_required
def relatorio_ferramentas():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        tickets_query = db.session.query(
            Ticket.tool,
            Ticket.status,
            func.count(Ticket.id).label('quantidade')
        )
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)
        atendimentos_por_ferramenta_status = tickets_query.group_by(Ticket.tool, Ticket.status).all()

        data_dict = {}
        for row in atendimentos_por_ferramenta_status:
            ferramenta = row.tool
            status = row.status
            quantidade = row.quantidade
            if ferramenta not in data_dict:
                data_dict[ferramenta] = {'ferramenta': ferramenta, 'abertos': 0, 'fechados': 0}
            if status == 'open':
                data_dict[ferramenta]['abertos'] += quantidade
            elif status == 'closed':
                data_dict[ferramenta]['fechados'] += quantidade
        data_list = list(data_dict.values())
        return jsonify(data_list)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
#######################################################################################################################################################
       
############################################      ROTA OK   ###########################################################################################################

@app.route('/relatorio_ferramentas_csr')
@login_required
@permission_required
def relatorio_ferramentas_csr():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        query = text("""
            SELECT 
                tool AS ferramenta,
                SUM(CASE WHEN status = 'open' THEN 1 ELSE 0 END) AS abertos,
                SUM(CASE WHEN status = 'closed' THEN 1 ELSE 0 END) AS fechados
            FROM 
                ticket
            WHERE 
                csr REGEXP '^[0-9]+$'
        """)
        params = {}
        if start_date:
            query = text(str(query) + " AND created_at >= :start_date")
            params['start_date'] = start_date
        if end_date:
            query = text(str(query) + " AND created_at <= :end_date")
            params['end_date'] = end_date
        atendimentos_por_ferramenta_status = db.session.execute(query, params).fetchall()
        data_list = [{'ferramenta': row.ferramenta, 'abertos': row.abertos, 'fechados': row.fechados} 
                     for row in atendimentos_por_ferramenta_status]
        return jsonify(data_list)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
    
#######################################################################################################################################################
############################################      ROTA OK   ###########################################################################################################

@app.route('/relatorio_prioridade')
@login_required
@permission_required
def relatorio_prioridade():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        tickets_query = db.session.query(
            Ticket.priority,
            Ticket.status,
            func.count(Ticket.id).label('quantidade')
        )
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)
        atendimentos_por_prioridade_status = tickets_query.group_by(Ticket.priority, Ticket.status).all()

        data_dict = {}
        for row in atendimentos_por_prioridade_status:
            prioridade = row.priority
            status = row.status
            quantidade = row.quantidade
            if prioridade not in data_dict:
                data_dict[prioridade] = {'prioridade': prioridade, 'abertos': 0, 'fechados': 0}
            if status == 'open':
                data_dict[prioridade]['abertos'] += quantidade
            elif status == 'closed':
                data_dict[prioridade]['fechados'] += quantidade
        data_list = list(data_dict.values())
        return jsonify(data_list)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
#######################################################################################################################################################
    
############################################      ROTA OK   ###########################################################################################################

@app.route('/relatorio_prioridade_pizza')
@login_required
@permission_required
def relatorio_prioridade_pizza():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        tickets_query = db.session.query(
            Ticket.priority,
            func.count(Ticket.id).label('quantidade')
        )
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)
        atendimentos_por_prioridade = tickets_query.group_by(Ticket.priority).all()

        labels = [row.priority for row in atendimentos_por_prioridade]
        values = [row.quantidade for row in atendimentos_por_prioridade]
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40'][:len(labels)]
        return jsonify({'labels': labels, 'values': values, 'colors': colors})
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
#######################################################################################################################################################

    
############################################      ROTA OK   ###########################################################################################################

@app.route('/relatorio_ferramentas_page')
@login_required
@permission_required
def relatorio_ferramentas_page():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        tickets_query = Ticket.query
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)
        tickets = tickets_query.all()

        atendimentos_por_ferramenta_status = db.session.query(
            Ticket.tool,
            Ticket.status,
            func.count(Ticket.id).label('quantidade')
        ).group_by(Ticket.tool, Ticket.status).all()

        ferramentas = []
        abertos = []
        fechados = []
        ferramenta_dict = {}

        for row in atendimentos_por_ferramenta_status:
            ferramenta = row.tool
            if ferramenta not in ferramenta_dict:
                ferramenta_dict[ferramenta] = {'abertos': 0, 'fechados': 0}
            if row.status == 'open':
                ferramenta_dict[ferramenta]['abertos'] += row.quantidade
            elif row.status == 'closed':
                ferramenta_dict[ferramenta]['fechados'] += row.quantidade

        for ferramenta, counts in ferramenta_dict.items():
            ferramentas.append(ferramenta)
            abertos.append(counts['abertos'])
            fechados.append(counts['fechados'])

        logger.debug(f"Ferramentas: {ferramentas}, Abertos: {abertos}, Fechados: {fechados}")
        return render_template('relatorio_ferramentas.html', 
                              ferramentas=ferramentas, 
                              abertos=abertos, 
                              fechados=fechados, 
                              tickets=tickets)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
    
#######################################################################################################################################################    
        
############################################      ROTA OK   ###########################################################################################################

@app.route('/plot_line')
@login_required
@permission_required
def plot_line():
    try:
        tickets = db.session.execute(
            text("SELECT csr, os, created_at FROM ticket WHERE created_at IS NOT NULL")
        ).fetchall()

        csr_data_dict = collections.defaultdict(int)
        os_data_dict = collections.defaultdict(int)
        for row in tickets:
            # CORREÇÃO: Converte a string de data para objeto datetime
            created_at_dt = parser.parse(row.created_at)
            date = created_at_dt.strftime('%Y-%m-%d')
            if row.csr:
                csr_data_dict[date] += 1
            if row.os:
                os_data_dict[date] += 1

        sorted_dates = sorted(set(csr_data_dict.keys()).union(os_data_dict.keys()))
        csr_cumulative_sum = []
        os_cumulative_sum = []
        csr_cumulative_value = 0
        os_cumulative_value = 0
        for date in sorted_dates:
            csr_cumulative_value += csr_data_dict[date]
            os_cumulative_value += os_data_dict[date]
            csr_cumulative_sum.append(csr_cumulative_value)
            os_cumulative_sum.append(os_cumulative_value)

        return jsonify({'labels': sorted_dates, 'csr_values': csr_cumulative_sum, 'os_values': os_cumulative_sum})
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


from datetime import datetime, timezone
from dateutil import parser
from zoneinfo import ZoneInfo

# Ajuste aqui se seus tempos no DB estão em outro fuso local
DEFAULT_LOCAL_TZ = ZoneInfo("America/Sao_Paulo")

def to_utc(dt, assume_tz=DEFAULT_LOCAL_TZ):
    """
    Recebe string ou datetime. Se for naive (sem tz), assume assume_tz
    (p.ex. America/Sao_Paulo). Retorna datetime timezone-aware em UTC.
    """
    if dt is None:
        return None
    if isinstance(dt, str):
        dt = parser.parse(dt)
    # dt é datetime agora
    if dt.tzinfo is None:
        # assume que timestamp sem tz é do fuso local (ex. Brasil)
        dt = dt.replace(tzinfo=assume_tz)
    # converte para UTC
    return dt.astimezone(timezone.utc)

def horas_entre(created_at, closed_at, assume_tz=ZoneInfo("America/Sao_Paulo")):
    """Retorna horas (float) entre created_at e closed_at, assumindo America/Sao_Paulo."""
    if created_at is None or closed_at is None:
        return 0.0
    if isinstance(created_at, str):
        created_at = parser.parse(created_at)
    if isinstance(closed_at, str):
        closed_at = parser.parse(closed_at)
    # Garante que ambas as datas estão em America/Sao_Paulo
    if created_at.tzinfo is None:
        created_at = created_at.replace(tzinfo=assume_tz)
    if closed_at.tzinfo is None:
        closed_at = closed_at.replace(tzinfo=assume_tz)
    delta = closed_at - created_at
    secs = max(delta.total_seconds(), 0.0)
    return secs / 3600.0

def format_delta_hours(hours):
    """Formata horas em 'Xd Yh Zm' para debug humano."""
    total_minutes = int(round(hours * 60))
    d, rem = divmod(total_minutes, 60*24)
    h, m = divmod(rem, 60)
    return f"{d}d {h}h {m}m"


from datetime import datetime
# ... (outros imports que já tinha)

def parse_paused_to_hours(paused):
    """
    Aceita paused como int/float (segundos) ou string 'HH:MM:SS' ou string numérica (segundos).
    Retorna horas (float).
    """
    if not paused:
        return 0.0
    try:
        # se já for numérico (segundos)
        if isinstance(paused, (int, float)):
            return float(paused) / 3600.0
        # se for string
        if isinstance(paused, str):
            # caso 'HH:MM:SS'
            if ':' in paused:
                parts = paused.split(':')
                if len(parts) == 3:
                    h, m, s = [float(p) for p in parts]
                    return (h * 3600 + m * 60 + s) / 3600.0
            # caso string que representa segundos "3600" ou "3600.0"
            try:
                return float(paused) / 3600.0
            except ValueError:
                return 0.0
    except Exception:
        return 0.0
    return 0.0

def horas_entre_raw(start, end):
    """Retorna diferença entre start e end em horas (sem subtrair pausas)."""
    if not start or not end:
        return 0.0
    if isinstance(start, str):
        try:
            start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S.%f')
        except ValueError:
            start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
    if isinstance(end, str):
        try:
            end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S.%f')
        except ValueError:
            end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
    delta = end - start
    return max(0.0, delta.total_seconds() / 3600.0)

#######################################################################################################################################################

@app.route('/relatorio_tempo_fechamento')
@login_required
@permission_required
def relatorio_tempo_fechamento_tool():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        query = """
            SELECT id, created_at, closed_at, tool, total_paused_duration
            FROM ticket 
            WHERE status = 'closed' AND closed_at IS NOT NULL
        """
        params = {}
        if start_date and end_date:
            query += " AND created_at >= :start_date AND closed_at <= :end_date"
            params['start_date'] = start_date
            params['end_date'] = end_date + ' 23:59:59'

        rows = db.session.execute(text(query), params).fetchall()

        if not rows:
            return jsonify({'labels': [], 'data_values': [], 'data_values_te': []})

        data = []
        for r in rows:
            raw_h = horas_entre_raw(r.created_at, r.closed_at)  # raw: closed_at - created_at
            paused_h = parse_paused_to_hours(r.total_paused_duration)  # horas de pausa
            effective_h = max(0.0, raw_h - paused_h)  # tempo efetivo subtraindo pausa (garante >=0)

            data.append({
                'tool': r.tool or 'N/A',
                'raw_hours': raw_h,
                'effective_hours': effective_h,
                'id': r.id
            })

        df = pd.DataFrame(data)

        tool_avg = df.groupby('tool', as_index=False).agg({
            'raw_hours': 'mean',
            'effective_hours': 'mean'
        })

        return jsonify({
            # data_values => média de (closed_at - created_at)  <-- MÉDIA DATA FECHAMENTO
            'labels': tool_avg['tool'].tolist(),
            'data_values': tool_avg['raw_hours'].round(2).tolist(),
            # data_values_te => média do tempo efetivo (subtraindo pausas) <-- MÉDIA TEMPO EFETIVO
            'data_values_te': tool_avg['effective_hours'].round(2).tolist()
        })
    except Exception as e:
        logger.error("Erro ao gerar relatório por ferramenta: %s\n%s", e, traceback.format_exc())
        return jsonify({"error": "Erro ao buscar dados."}), 500



@app.route('/plot_line_closed')
@login_required
@permission_required
def plot_line_closed():
    try:
        tickets = db.session.execute(
            text("SELECT csr, os, closed_at FROM ticket WHERE status = 'closed' AND closed_at IS NOT NULL")
        ).fetchall()

        csr_data_dict = collections.defaultdict(int)
        os_data_dict = collections.defaultdict(int)
        for row in tickets:
            if row.closed_at:
                closed_at_dt = parser.parse(row.closed_at)
                date = closed_at_dt.strftime('%Y-%m-%d')
                if row.csr:
                    csr_data_dict[date] += 1
                if row.os:
                    os_data_dict[date] += 1
       
        sorted_dates = sorted(set(csr_data_dict.keys()).union(os_data_dict.keys()))
        csr_cumulative_sum = []
        os_cumulative_sum = []
        csr_cumulative_value = 0
        os_cumulative_value = 0
        for date in sorted_dates:
            csr_cumulative_value += csr_data_dict[date]
            os_cumulative_value += os_data_dict[date]
            csr_cumulative_sum.append(csr_cumulative_value)
            os_cumulative_sum.append(os_cumulative_value) # CORREÇÃO: era 'os_cumulative_sum.append(os_cumulative_sum)'
           
        return jsonify({'labels': sorted_dates, 'csr_values': csr_cumulative_sum, 'os_values': os_cumulative_sum})
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500



@app.route('/sla_chart_geral')
@login_required
@permission_required
def sla_chart_geral():
    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        results = db.session.execute(
            text("""
                SELECT tool,
                       CASE
                           WHEN closed_at IS NULL THEN JULIANDAY(:current_date) - JULIANDAY(created_at)
                           ELSE JULIANDAY(closed_at) - JULIANDAY(created_at)
                       END AS sla_days
                FROM ticket
            """),
            {"current_date": current_date}
        ).fetchall()

        sla_data = {}
        for tool, sla_days in results:
            if tool not in sla_data:
                sla_data[tool] = {"Dentro do SLA": 0, "Fora do SLA": 0}
            if sla_days <= 2:
                sla_data[tool]["Dentro do SLA"] += 1
            else:
                sla_data[tool]["Fora do SLA"] += 1

        tools = list(sla_data.keys())
        dentro_sla = [sla_data[tool]["Dentro do SLA"] for tool in tools]
        fora_sla = [sla_data[tool]["Fora do SLA"] for tool in tools]
        return jsonify({"tools": tools, "dentro_sla": dentro_sla, "fora_sla": fora_sla})
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/ticket_counts_chart')
@login_required
@permission_required
def ticket_counts_chart():
    try:
        total_id = db.session.query(func.count()).filter(Ticket.id.cast(db.String).op('GLOB')('[0-9]*')).scalar()
        total_csr = db.session.query(func.count()).filter(Ticket.csr.cast(db.String).op('GLOB')('[0-9]*')).scalar()
        dif_idxcsr = total_id - total_csr
        data = {
            "labels": ["IDS-Brasil", "QDT total Ocorrência", "IDS-IT CSR"],
            "values": [dif_idxcsr, total_id, total_csr],
            "colors": ["#FF6384", "#36A2EB", "#FFCE56"]
        }
        return jsonify(data)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500



@app.route('/acompanhamento')
@login_required
@permission_required
def acompanhamento():
    try:
        file_path = 'E:/Python/Abertura_de_Atendimento/templates/tickets.xls'
        if not os.path.exists(file_path):
            flash('Arquivo tickets.xls não encontrado.', 'error')
            return redirect(url_for('index'))
        df = pd.read_excel(file_path)
        expected_columns = [
            'Ticket Id', 'Subject', 'Status', 'Priority', 'Type',
            'Created Time', 'Last Updated Time', 'Software Product',
            'TR Workitem ID', 'CR Workitem ID', 'TRCR Status'
        ]
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            flash(f'Colunas faltando no arquivo: {", ".join(missing_columns)}', 'warning')
        df = df[expected_columns] if all(col in df.columns for col in expected_columns) else df[list(set(expected_columns) & set(df.columns))]
        tickets = df.to_dict(orient='records')
        return render_template('acompanhamento.html', tickets=tickets)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        flash(f'Erro ao carregar o arquivo: {str(e)}', 'error')
        return redirect(url_for('index'))



@app.route('/csr_data')
@login_required
@permission_required
def csr_data():
    try:
        conn = get_csr_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT LOWER(status) AS status, COUNT(*) FROM tickets GROUP BY LOWER(status)")
        rows = cursor.fetchall()
        conn.close()
        labels = [row[0] for row in rows]
        quantities = [row[1] for row in rows]
        logger.debug(f"Dados do gráfico: labels={labels}, quantities={quantities}")
        return jsonify({'labels': labels, 'quantities': quantities})
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/get_stats')
@login_required
def get_stats():
    try:
        total_tickets = Ticket.query.count()
        open_tickets = Ticket.query.filter_by(status='open').count()
        closed_tickets = Ticket.query.filter_by(status='closed').count()
        
        # Estatísticas para o dia e semana
        today = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
        start_of_week = today - timedelta(days=today.weekday())
        closed_today = Ticket.query.filter(
            Ticket.closed_at >= datetime.combine(today, datetime.min.time(), tzinfo=ZoneInfo("America/Sao_Paulo")),
            Ticket.closed_at < datetime.combine(today, datetime.max.time(), tzinfo=ZoneInfo("America/Sao_Paulo"))
        ).count()
        
        closed_week = Ticket.query.filter(
            Ticket.closed_at >= datetime.combine(start_of_week, datetime.min.time(), tzinfo=ZoneInfo("America/Sao_Paulo"))
        ).count()
        return jsonify({
            'total_tickets': total_tickets,
            'open_tickets': open_tickets,
            'closed_tickets': closed_tickets,
            'closed_today': closed_today,
            'closed_week': closed_week
        })
    except Exception as e:
        logger.error(f"Erro ao buscar estatísticas: {e}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


# Defina esta função no topo do seu arquivo
def gerar_os_aleatoria():
    """Gera um número de OS aleatório com 6 caracteres (letras maiúsculas e números)."""
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choices(caracteres, k=6))

@app.route('/solicitar_os/<int:ticket_id>', methods=['POST'])
@login_required
def solicitar_os(ticket_id):
    try:
        # Encontre o ticket no seu banco de dados
        ticket = db.session.get(Ticket, ticket_id)
        
        # Verifique se o ticket existe
        if not ticket:
            flash('Ticket não encontrado.', 'error')
            return redirect(url_for('index'))
            
        # Verifique se o campo AUX2 (OS) já está preenchido
        if ticket.AUX2 and ticket.AUX2.strip():
            flash('Este ticket já possui um número de OS.', 'warning')
            return redirect(url_for('ticket_details', ticket_id=ticket.id))
            
        # Gere a nova OS
        nova_os = gerar_os_aleatoria()
        
        # Salva o novo número APENAS no campo 'AUX2'
        ticket.AUX2 = nova_os
        
        # Salve as alterações no banco de dados
        db.session.commit()
        
        flash(f'OS {nova_os} solicitada e salva para o ticket #{ticket_id}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f"Ocorreu um erro ao solicitar a OS: {e}", 'error')
        
    return redirect(url_for('ticket_details', ticket_id=ticket.id))


@app.route('/', methods=['GET', 'POST'])
@login_required
@permission_required
def index():
    try:
        # ===================================================================
        # ##           CÁLCULO DAS ESTATÍSTICAS (FEITO PRIMEIRO)         ##
        # ===================================================================
        # Estes cálculos são feitos na tabela inteira, ANTES de qualquer filtro.
        stats = {
            'total_tickets': Ticket.query.count(),
            'total_abertos': Ticket.query.filter_by(status='open').count(),
            'total_pausados': Ticket.query.filter_by(status='open', status_sla='paused').count(),
            'total_fechados': Ticket.query.filter_by(status='closed').count()
        }
        # ===================================================================

        # ===================================================================
        # ##           LÓGICA DE FILTRO PARA A TABELA EXIBIDA            ##
        # ===================================================================
        # Pega os argumentos da URL para os filtros
        username_filter = request.args.get('username', '')
        tool_filter = request.args.get('tool', '')
        # Adicione outros filtros que você possa ter aqui

        # Começa com uma consulta base
        tickets_query = Ticket.query

        # Aplica os filtros se eles existirem
        if username_filter:
            tickets_query = tickets_query.filter(Ticket.username.ilike(f'%{username_filter}%'))
        if tool_filter:
            tickets_query = tickets_query.filter(Ticket.tool == tool_filter)
        
        # Ordena e executa a consulta FINAL para a tabela
        tickets_para_exibir = tickets_query.order_by(Ticket.created_at.desc()).all()
        # ===================================================================
        # ===================================================================
        # ##           CÁLCULO DA COLUNA TC/CR                           ##
        # ===================================================================
        file_path = 'E:/Python/Abertura_de_Atendimento/templates/tickets.xls'
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, dtype={'Ticket Id': str})
            tc_cr_map = {}
            for _, row in df.iterrows():
                ticket_id = row.get('Ticket Id')
                if ticket_id is not None:
                    ticket_id_str = str(ticket_id).strip()
                    tr = row.get('TR Workitem ID')
                    cr = row.get('CR Workitem ID')
                    if pd.notna(tr) or pd.notna(cr):
                        tc_cr_map[ticket_id_str] = "SIM"
            logger.debug(f"Loaded TC/CR map: {tc_cr_map}")
            for ticket in tickets_para_exibir:
                csr_str = str(ticket.csr).strip() if ticket.csr else ""
                if csr_str in tc_cr_map:
                    ticket.tc_cr = tc_cr_map[csr_str]
                    logger.debug(f"Match found for CSR {csr_str} in ticket {ticket.id}")
                else:
                    ticket.tc_cr = ""
                    logger.debug(f"No match for CSR {csr_str} in ticket {ticket.id}")
        else:
            for ticket in tickets_para_exibir:
                ticket.tc_cr = ""
            logger.warning("Arquivo tickets.xls não encontrado para cálculo de TC/CR.")
        # ===================================================================

        return render_template(
            'tickets_list.html', 
            tickets=tickets_para_exibir, # Passa a lista filtrada para a tabela
            stats=stats                  # Passa as estatísticas totais para os cards
        )
        
    except Exception as e:
        logger.error(f"Erro na página inicial: {e}\n{traceback.format_exc()}")
        # Garante que a página não quebre se houver erro
        return render_template('tickets_list.html', tickets=[], stats={})




@app.route('/edit_ticket/<int:ticket_id>', methods=['GET', 'POST'])
@login_required
@permission_required
def edit_ticket(ticket_id):
    try:
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket não encontrado.', 'error')
            return redirect(url_for('index'))

        if request.method == 'POST':
            # --- INÍCIO DA SEÇÃO DE DEPURAÇÃO ---
            print("--- DADOS RECEBIDOS DO FORMULÁRIO ---")
            print(request.form)
            valor_ferramenta_recebido = request.form.get('tool')
            print(f"Valor recebido para 'tool': {valor_ferramenta_recebido}")
            # --- FIM DA SEÇÃO DE DEPURAÇÃO ---

            # Atualiza os campos do ticket
            ticket.username = request.form.get('username', ticket.username)
            ticket.tool = valor_ferramenta_recebido or ticket.tool # Usamos a variável para garantir
            ticket.subject = request.form.get('subject', ticket.subject)
            ticket.description = request.form.get('description', ticket.description)
            ticket.csr = request.form.get('csr', ticket.csr)
            ticket.priority = request.form.get('priority', ticket.priority)
            
            db.session.commit()
            flash('Ticket atualizado com sucesso!', 'success')
            return redirect(url_for('ticket_details', ticket_id=ticket.id))

        # Se o método for GET, apenas renderiza a página
        return render_template('edit_ticket.html', ticket=ticket) # Certifique-se que o template está correto

    except Exception as e:
        logger.error(f"Erro em edit_ticket: {e}\n{traceback.format_exc()}")
        flash('Ocorreu um erro ao editar o ticket.', 'error')
        return redirect(url_for('ticket_details', ticket_id=ticket_id))


@app.route('/ticket/<int:ticket_id>', methods=['GET', 'POST'])
@login_required

def ticket_details(ticket_id):
    try:
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket não encontrado.', 'error')
            return redirect(url_for('index'))
        comments = Comment.query.filter_by(ticket_id=ticket_id).order_by(Comment.created_at.desc()).all()
        return render_template('ticket_details.html', ticket=ticket, comments=comments)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/ticket/<int:ticket_id>/close', methods=['POST'])
@login_required

def close_ticket(ticket_id):
    try:
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket não encontrado.', 'error')
            return redirect(url_for('index'))
        ticket.status = 'closed'
        ticket.closed_at = datetime.now(ZoneInfo("America/Sao_Paulo"))  # Alterado de utcnow
        db.session.commit()
        flash('Ticket fechado com sucesso.', 'success')
        return redirect(url_for('ticket_details', ticket_id=ticket.id))
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/ticket/<int:ticket_id>/comment', methods=['POST'])
@login_required
@permission_required
def add_comment(ticket_id):
    try:
        username = request.form.get('username')
        comment_text = request.form.get('comment')
        files = request.files.getlist('files')

        if username and comment_text:
            comment = Comment(
                ticket_id=ticket_id,
                username=username,
                text=comment_text,
                created_at=datetime.utcnow()
            )
            db.session.add(comment)
            db.session.commit()

            if save_comment_to_excel(str(ticket_id), username, comment_text):
                flash('Comentário adicionado e salvo no Excel com sucesso.', 'success')
            else:
                flash('Comentário adicionado, mas não foi possível salvá-lo no Excel.', 'warning')

            for file in files:
                if file:
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    comment_file = CommentFile(comment_id=comment.id, filename=filename)
                    db.session.add(comment_file)

            db.session.commit()
        return redirect(url_for('ticket_details', ticket_id=ticket_id))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        flash('Ocorreu um erro ao adicionar o comentário e anexos.', 'error')
        return redirect(url_for('ticket_details', ticket_id=ticket_id))
    finally:
        db.session.remove()

def formatar_tempo_em_texto(minutos):
    if minutos < 60:
        return f"{minutos} minutes"
    elif minutos < 1440:
        horas = minutos // 60
        return f"{horas} hours"
    elif minutos < 43200:
        dias = minutos // 1440
        return f"{dias} days"
    elif minutos < 525600:
        meses = minutos // 43200
        return f"{meses} months"
    else:
        anos = minutos // 525600
        return f"{anos} years"

def extrair_tempo_em_minutos(status):
    try:
        match = re.search(r'(\d+|a)\s*(minute|minutes|hour|hours|day|days|month|months|year|years)', status)
        if match:
            time_value = match.group(1)
            time_unit = match.group(2)
            time_value = 1 if time_value == "a" else int(time_value)
            if 'minute' in time_unit:
                return time_value
            elif 'hour' in time_unit:
                return time_value * 60
            elif 'day' in time_unit:
                return time_value * 1440
            elif 'month' in time_unit:
                return time_value * 43200
            elif 'year' in time_unit:
                return time_value * 525600
        return float('inf')
    except Exception as e:
        logger.error(f"Erro ao extrair tempo: {e}")
        return float('inf')


@app.route('/ticket/<int:ticket_id>/pause', methods=['POST'])
@login_required
@permission_required
def pause_ticket(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if ticket and ticket.status == 'open' and ticket.status_sla == 'running':
        ticket.status_sla = 'paused'
        ticket.paused_at = datetime.now(ZoneInfo("America/Sao_Paulo"))  # Alterado de utcn
        db.session.commit()
        flash(f'O ticket {ticket.id} foi pausado.', 'success')
    else:
        flash('Não foi possível pausar o ticket.', 'error')
    return redirect(url_for('index'))

@app.route('/ticket/<int:ticket_id>/resume', methods=['POST'])
@login_required
@permission_required
def resume_ticket(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if ticket and ticket.status == 'open' and ticket.status_sla == 'paused':
        if ticket.paused_at:
            pause_duration = (datetime.now() - ticket.paused_at).total_seconds()
            ticket.total_paused_duration += int(pause_duration)
        
        ticket.status_sla = 'running'
        ticket.paused_at = None
        db.session.commit()
        flash(f'O ticket {ticket.id} foi retomado.', 'success')
    else:
        flash('Não foi possível retomar o ticket.', 'error')
    return redirect(url_for('index'))


# Em APP_OFICIAL2.Py, adicione esta nova rota

@app.route('/relatorio_sla')
@login_required
@permission_required
def relatorio_sla():
    try:
        tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()
        
        report_data = []
        for ticket in tickets:
            # Duração total desde a criação até agora (ou até o fechamento)
            from datetime import datetime, timezone

            end_time = ticket.closed_at or datetime.now(timezone.utc)
            total_duration = end_time - ticket.created_at
            
            # Duração total em que o ticket esteve pausado
            paused_duration = timedelta(seconds=ticket.total_paused_duration)
            
            # Duração efetiva (Total - Pausado)
            effective_duration = total_duration - paused_duration

            report_data.append({
                'id': ticket.id,
                'subject': ticket.subject,
                'status': ticket.status,
                'total_duration_str': str(total_duration).split('.')[0], # Formata para 'dias, H:M:S'
                'paused_duration_str': str(paused_duration).split('.')[0],
                'effective_duration_str': str(effective_duration).split('.')[0]
            })
            
        return render_template('relatorio_sla.html', report_data=report_data)
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de SLA: {e}\n{traceback.format_exc()}")
        flash('Ocorreu um erro ao gerar o relatório de SLA.', 'error')
        return redirect(url_for('index'))



import sqlite3
import os

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "instance", "app.db")

def get_db_connection():
    db_path = r"E:\Python\Abertura_de_Atendimento\instance\app.db"
    print(">>> USANDO BANCO:", db_path)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

import sqlite3

DB_PATH = r"E:\Python\Abertura_de_Atendimento\instance\app.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def buscar_comentario(ticket_url):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT comentario FROM comentarios WHERE ticket_url = ?",
        (ticket_url,)
    )
    row = cur.fetchone()
    conn.close()
    return row["comentario"] if row else ""

from datetime import datetime

@app.route("/salvar_comentario", methods=["POST"])
@login_required
@permission_required
def salvar_comentario():
    try:
        data = request.get_json()
        ticket_url = data.get("ticket_url")
        comentario = data.get("comentario", "").strip()

        if not ticket_url:
            return jsonify({"success": False, "error": "URL inválida"})

        # 🔹 Data e hora no formato desejado
        agora = datetime.now().strftime("%d/%m/%Y %H:%M")

        # 🔹 Comentário final
        comentario_formatado = f"{agora} - {comentario}"

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO comentarios (ticket_url, comentario)
            VALUES (?, ?)
            ON CONFLICT(ticket_url)
            DO UPDATE SET comentario = excluded.comentario
        """, (ticket_url, comentario_formatado))

        conn.commit()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        logger.error(e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500



@app.route('/Fresdesk')
@login_required
@permission_required
def Fresdesk():
    try:
        # ==========================
        # 1. LÊ O EXCEL (FONTE DOS TICKETS)
        # ==========================
        excel_path = r"E:/Python/Coleta-Site/exportado_formatado.xlsx"

        df = pd.read_excel(excel_path)

        if df.empty:
            return "Erro: Nenhum ticket encontrado."

        if "Última atualização" not in df.columns or "URL" not in df.columns:
            return "Erro: Colunas obrigatórias não encontradas no Excel."

        # ==========================
        # 2. TRATAMENTO DE DATA
        # ==========================
        df["Última atualização"] = pd.to_datetime(
            df["Última atualização"],
            errors="coerce",
            dayfirst=True
        )

        df = df[df["Última atualização"].notna()]

        agora = pd.Timestamp.now()

        df["Ultima Atualização Minutos"] = (
            agora - df["Última atualização"]
        ).dt.total_seconds() / 60

        df["ultima_atualizacao_texto"] = df[
            "Ultima Atualização Minutos"
        ].apply(formatar_tempo_em_texto)

        df["ultima_atualizacao_data"] = df[
            "Última atualização"
        ].dt.strftime("%d/%m/%Y %H:%M")

        # ==========================
        # 3. BUSCA COMENTÁRIOS DO BANCO (ESTÁVEIS)
        # ==========================
        conn = get_db()
        comentarios_df = pd.read_sql(
            "SELECT ticket_url, comentario FROM comentarios",
            conn
        )
        conn.close()

        # ==========================
        # 4. MERGE EXCEL + BANCO
        # ==========================
        df = df.merge(
            comentarios_df,
            how="left",
            left_on="URL",
            right_on="ticket_url"
        )

        df["comentario"] = df["comentario"].fillna("")

        # ==========================
        # 5. ORDENAÇÃO FINAL
        # ==========================
        df = df.sort_values(by="Última atualização", ascending=False)

        # ==========================
        # 6. RENDERIZA
        # ==========================
        return render_template(
            "export_csr.html",
            data=df.to_dict(orient="records")
        )

    except Exception as e:
        logger.error(e, exc_info=True)
        return f"Erro ao processar: {e}"



def salvar_ticket(ticket_id, url, status, ultima_atualizacao, comentario):
    conn = sqlite3.connect("freshdesk.db")
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO tickets (ticket_id, url, status, ultima_atualizacao, comentario, updated_at)
    VALUES (?, ?, ?, ?, ?, datetime('now'))
    ON CONFLICT(ticket_id) DO UPDATE SET
        status = excluded.status,
        ultima_atualizacao = excluded.ultima_atualizacao,
        updated_at = datetime('now'),
        comentario = CASE
            WHEN tickets.comentario IS NULL OR tickets.comentario = ''
            THEN excluded.comentario
            ELSE tickets.comentario
        END
    """, (ticket_id, url, status, ultima_atualizacao, comentario))

    conn.commit()
    conn.close()


@app.route('/ticket')
def ticket_detail():
    csr = request.args.get('csr')
    if not csr:
        return "Erro: parâmetro 'csr' não informado.", 400

    # Ler planilha
    import pandas as pd
    df = pd.read_excel("E:/Python/Coleta-Site/exportado_formatado.xlsx")

    # Procurar o CSR correspondente
    row = df.loc[df['URL'] == csr].to_dict(orient='records')
    if not row:
        return f"CSR {csr} não encontrado.", 404

    return render_template('ticket_detail9.html', data=row[0])


@app.route('/register', methods=['GET', 'POST'])
def register():
    try:
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            if User.query.filter_by(username=username).first():
                flash('Usuário já existe. Escolha outro nome.', 'error')
                return redirect(url_for('register'))
            new_user = User(username=username, is_admin=False)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            flash('Usuário criado com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
        return render_template('register.html')
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

   

###########################   Indicador de SLA  Relatório Detalhado de SLA

# Configurar logging para depuração
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Função format_time_diff (reutilizada do /relatorio_tempo_fechamento)
def format_time_diff(start, end, paused_duration=0):
    """
    Formata a diferença de tempo entre start e end no formato 'Xd Yh Zm', subtraindo paused_duration (em segundos).
    """
    if not start or not end:
        return '0d 00h 00m'
    # Convert strings to datetime if necessary, handling microseconds
    if isinstance(start, str):
        try:
            start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S.%f')
        except ValueError:
            start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
    if isinstance(end, str):
        try:
            end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S.%f')
        except ValueError:
            end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
    delta = end - start
    total_seconds = delta.total_seconds()
    # Subtrair a duração das pausas, se disponível
    if paused_duration:
        total_seconds -= paused_duration
    total_seconds = max(0, total_seconds)  # Garante valor não negativo
    days = int(total_seconds // (24 * 3600))
    hours = int((total_seconds % (24 * 3600)) // 3600)
    minutes = int((total_seconds % 3600) // 60)
    time_diff = f"{days}d {hours:02d}h {minutes:02d}m"
    logger.debug(f"format_time_diff: start={start}, end={end}, paused_duration={paused_duration}, time_diff={time_diff}")
    return time_diff

@app.route('/relatorio_sla_detalhado')
@login_required
def relatorio_sla_detalhado():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        tickets_query = Ticket.query
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            # Ajustar end_date para incluir o dia inteiro
            end_date = f"{end_date} 23:59:59"
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)

        tickets = tickets_query.order_by(Ticket.created_at.desc()).all()
        data = []

        # ==============================
        # TABELA DE REGRAS SLA
        # ==============================
        REGRAS_SLA = {
            'AIMPORTAL':        {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'AIRSPACE DESIGNER':{'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'AERODB':           {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'ETOD':             {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'GFEAMAN':          {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'ICE':              {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'PLX':              {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'WEPUB':            {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'CRONOS':           {'Crítica': 1, 'Grave': 2, 'Moderada': 4, 'Baixo': float('inf')},
            'DDP':              {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'FPDAM':            {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'PROJECT MANAGER':  {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},        
        }
        DEFAULT_REGRAS = {'Crítica': float('inf'), 'Grave': float('inf'),
                          'Moderada': float('inf'), 'Baixo': float('inf')}

        for ticket in tickets:
            # --- Cálculo do Tempo Efetivo ---
            end_time = ticket.closed_at or datetime.now()
            # Usar format_time_diff para consistência
            effective_duration_str = format_time_diff(ticket.created_at, end_time, ticket.total_paused_duration or 0)

            # ================================================================
            # LÓGICA DE SLA ATENDIDO
            # ================================================================
            sla_atendido = 'Em Andamento'
            sla_calc = None  # Usado para o gráfico

            if ticket.status.lower() == 'closed':
                tool_upper = (ticket.tool or '').strip().upper()
                prioridade_key = (ticket.priority or '').strip().title()

                criterios = REGRAS_SLA.get(tool_upper, DEFAULT_REGRAS)
                max_days = criterios.get(prioridade_key, float('inf'))

                # Calcular effective_duration para comparação com SLA
                total_duration = end_time - ticket.created_at
                paused_duration = timedelta(seconds=ticket.total_paused_duration or 0)
                effective_duration = total_duration - paused_duration

                if max_days == float('inf'):
                    sla_atendido = 'Aguardando Versão'  # Mostra na tabela
                    sla_calc = 'Sim'  # Conta como cumprido no gráfico
                elif effective_duration.days <= max_days:
                    sla_atendido = 'Sim'
                    sla_calc = 'Sim'
                else:
                    sla_atendido = 'Não'
                    sla_calc = 'Não'
            else:
                sla_calc = None

            # Adicionar dados ao resultado
            data.append({
                'ticket_id': ticket.id,
                'os': ticket.os or 'N/A',
                'ods': ticket.AUX2 or 'N/A',
                'csr': ticket.csr or 'N/A',
                'username': ticket.username or 'N/A',
                'tool': ticket.tool or 'N/A',
                'priority': ticket.priority or 'N/A',
                'subject': ticket.subject or 'N/A',
                'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else 'N/A',
                'closed_at': ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else 'N/A',
                'status': ticket.status or 'N/A',
                'status_sla': 'Pausado' if ticket.status_sla == 'paused' else 'Em Andamento',
                'time_diff': effective_duration_str,
                'sla_atendido': sla_atendido,
                'sla_calc': sla_calc
            })
            logger.debug(f"Ticket ID={ticket.id}, tool={ticket.tool}, time_diff={effective_duration_str}, sla_atendido={sla_atendido}, sla_calc={sla_calc}")

        logger.debug(f"Dados retornados: {data}")
        return jsonify(data)
    except Exception as e:
        logger.error(f"Erro ao gerar relatório detalhado de SLA: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

  

from datetime import datetime, timedelta, timezone


################################################   SLA por Ferramenta  ################################################
@app.route('/relatorio_sla_ferramentas_hodometro')
@login_required
def relatorio_sla_ferramentas_hodometro():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')

        tickets_query = Ticket.query
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            tickets_query = tickets_query.filter(Ticket.created_at <= end_date)

        tickets = tickets_query.all()

        # Regras SLA
        REGRAS_SLA = {
            'AIMPORTAL':        {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'AIRSPACE DESIGNER':{'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'AERODB':           {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'ETOD':             {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'GFEAMAN':          {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'ICE':              {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'PLX':              {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'WEPUB':            {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'CRONOS':           {'Crítica': 1, 'Grave': 2, 'Moderada': 4, 'Baixo': float('inf')},
            'DDP':              {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'FPDAM':            {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'PROJECT MANAGER':  {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')}, 
        }
        DEFAULT_REGRAS = {'Crítica': float('inf'), 'Grave': float('inf'),
                          'Moderada': float('inf'), 'Baixo': float('inf')}

        ferramentas_data = {}

        for ticket in tickets:
            tool_upper = (ticket.tool or "N/A").strip().upper()
            if tool_upper not in ferramentas_data:
                ferramentas_data[tool_upper] = {"total": 0, "atendidos": 0}

            # Normaliza datas (sempre timezone-aware em UTC)
            end_time = ticket.closed_at or datetime.now(timezone.utc)
            if end_time.tzinfo is None:
                end_time = end_time.replace(tzinfo=timezone.utc)

            created_at = ticket.created_at
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)

            total_duration = end_time - created_at
            paused_duration = timedelta(seconds=ticket.total_paused_duration or 0)
            effective_duration = total_duration - paused_duration

            prioridade_key = (ticket.priority or "").strip().capitalize()
            criterios = REGRAS_SLA.get(tool_upper, DEFAULT_REGRAS)
            max_days = criterios.get(prioridade_key, float("inf"))

            ferramentas_data[tool_upper]["total"] += 1

            if ticket.status == "closed":
                if max_days == float("inf") or effective_duration.days <= max_days:
                    ferramentas_data[tool_upper]["atendidos"] += 1

        result_data = []
        for tool, valores in ferramentas_data.items():
            total = valores["total"]
            atendidos = valores["atendidos"]
            nao_atendidos = total - atendidos
            sla_percent = (atendidos / total * 100) if total > 0 else 0
            result_data.append({
                "ferramenta": tool,
                "total": total,
                "atendidos": atendidos,
                "nao_atendidos": nao_atendidos,
                "sla_percent": round(sla_percent, 1)
            })

        return jsonify(result_data)

    except Exception as e:
        logger.error(f"Erro ao gerar SLA por ferramentas: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


################################################################################################################################################

 
# Definir REGRAS_SLA e DEFAULT_REGRAS (se não estiverem em outro módulo)
REGRAS_SLA = {
    'FERRAMENTA1': {
        'Alta': 2,
        'Média': 5,
        'Baixa': 7
    },
    'FERRAMENTA2': {
        'Alta': 1,
        'Média': 3,
        'Baixa': 5
    },
    # Adicione outras ferramentas conforme necessário
}

DEFAULT_REGRAS = {
    'Alta': 3,
    'Média': 7,
    'Baixa': 10
}

def format_time_diff(start, end, paused_duration):
    """
    Função para calcular e formatar a diferença de tempo entre start e end, considerando paused_duration.
    Retorna a diferença formatada como string (ex.: '14d 00h 00m').
    """
    try:
        total_duration = end - start
        paused_duration = timedelta(seconds=paused_duration or 0)
        effective_duration = total_duration - paused_duration
        
        days = effective_duration.days
        seconds = effective_duration.seconds
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        
        return f"{days}d {hours:02d}h {minutes:02d}m"
    except Exception as e:
        logger.error(f"Erro em format_time_diff: {str(e)}")
        return "N/A"  
  
  
@app.route('/update_ticket_dates', methods=['POST'])
@login_required
def update_ticket_dates():
    try:
        data = request.get_json()
        ticket_id = data.get('ticket_id')
        field = data.get('field')
        value = data.get('value')

        if not ticket_id or not field:
            return jsonify({"success": False, "error": "Ticket ID e campo são obrigatórios"}), 400

        if field not in ['created_at', 'closed_at']:
            return jsonify({"success": False, "error": "Campo inválido"}), 400

        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket não encontrado"}), 404

        # Convert the value to a datetime object if provided
        if value:
            try:
                # Parse the date in 'YYYY-MM-DD HH:MM' format
                date_value = datetime.strptime(value, '%Y-%m-%d %H:%M')
                setattr(ticket, field, date_value)
            except ValueError:
                return jsonify({"success": False, "error": "Formato de data inválido. Use YYYY-MM-DD HH:MM"}), 400
        else:
            # Allow clearing the closed_at field
            if field == 'closed_at':
                setattr(ticket, field, None)
            else:
                return jsonify({"success": False, "error": "A data de criação não pode ser vazia"}), 400

        # Update the database
        db.session.commit()

        # Recalculate time_diff and sla_atendido if necessary
        end_time = ticket.closed_at or datetime.now()
        effective_duration_str = format_time_diff(ticket.created_at, end_time, ticket.total_paused_duration or 0)

        # Recalculate SLA
        sla_atendido = 'Em Andamento'
        sla_calc = None
        if ticket.status.lower() == 'closed':
            tool_upper = (ticket.tool or '').strip().upper()
            prioridade_key = (ticket.priority or '').strip().title()
            criterios = REGRAS_SLA.get(tool_upper, DEFAULT_REGRAS)
            max_days = criterios.get(prioridade_key, float('inf'))

            total_duration = end_time - ticket.created_at
            paused_duration = timedelta(seconds=ticket.total_paused_duration or 0)
            effective_duration = total_duration - paused_duration

            if max_days == float('inf'):
                sla_atendido = 'Aguardando Versão'
                sla_calc = 'Sim'
            elif effective_duration.days <= max_days:
                sla_atendido = 'Sim'
                sla_calc = 'Sim'
            else:
                sla_atendido = 'Não'
                sla_calc = 'Não'

        return jsonify({
            "success": True,
            "ticket_id": ticket.id,
            "created_at": ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else 'N/A',
            "closed_at": ticket.closed_at.strftime('%Y-%m-%d %H:%M') if ticket.closed_at else 'N/A',
            "time_diff": effective_duration_str,
            "sla_atendido": sla_atendido,
            "sla_calc": sla_calc
        })
    except Exception as e:
        logger.error(f"Erro ao atualizar datas do ticket: {str(e)}\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
  
          

#@app.route('/relatorio_sla_real_ferramentas')
@login_required
@permission_required
def relatorio_sla_real_ferramentas():
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        # Query base dos tickets
        tickets_query = Ticket.query
        if start_date:
            tickets_query = tickets_query.filter(Ticket.created_at >= start_date)
        if end_date:
            # Adiciona +1 dia ao end_date para incluir o dia inteiro
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            tickets_query = tickets_query.filter(Ticket.created_at < end_date_dt)

        tickets = tickets_query.all()
        
        # Regras de SLA por ferramenta (as mesmas que você já definiu)
        REGRAS_SLA = {
            'AIMPORTAL':        {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'AIRSPACE DESIGNER':{'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'AERODB':           {'Crítica': 3, 'Grave': 4, 'Moderada': 6, 'Baixo': float('inf')},
            'ETOD':             {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'GFEAMAN':          {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'ICE':              {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'PLX':              {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'WEPUB':            {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'CRONOS':           {'crítica': 1, 'grave': 2, 'moderada': 4, 'baixo': float('inf')},
            'DDP':              {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'FPDAM':            {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
            'PROJECT MANAGER':  {'crítica': 3, 'grave': 4, 'moderada': 6, 'baixo': float('inf')},
        }
        DEFAULT_REGRAS = {'crítica': float('inf'), 'grave': float('inf'),
                          'moderada': float('inf'), 'baixo': float('inf')}

        # Dicionário para agrupar os dados por ferramenta
        ferramentas_data = {}
        
        for ticket in tickets:
            tool_name = ticket.tool or 'N/A'
            
            # Inicializa a ferramenta no dicionário se for a primeira vez
            if tool_name not in ferramentas_data:
                ferramentas_data[tool_name] = {
                    'total': 0,
                    'atendidos': 0
                }
            
            # --- Lógica de cálculo de SLA para cada ticket ---
            end_time = ticket.closed_at or datetime.now(timezone.utc)
            # Usamos a função to_utc para garantir a consistência
            total_duration = to_utc(end_time) - to_utc(ticket.created_at)
            paused_duration = timedelta(seconds=ticket.total_paused_duration)
            effective_duration = total_duration - paused_duration
            
            tool_upper = (ticket.tool or '').strip().upper()
            prioridade_key = (ticket.priority or '').strip().lower()
            
            criterios = REGRAS_SLA.get(tool_upper, DEFAULT_REGRAS)
            max_days = criterios.get(prioridade_key, float('inf'))
            
            effective_days = effective_duration.total_seconds() / 86400.0
            
            # Incrementa o total de tickets para a ferramenta
            ferramentas_data[tool_name]['total'] += 1
            
            # Considera como "atendido" se estiver dentro do prazo ou se for "Aguardando nova versão"
            if max_days == float('inf') or effective_days <= max_days:
                ferramentas_data[tool_name]['atendidos'] += 1
        
        # --- Prepara os dados para enviar como JSON ---
        result_data = []
        for tool_name, data in ferramentas_data.items():
            if data['total'] > 0:
                sla_percent = (data['atendidos'] / data['total']) * 100
                
                # Determina a cor baseada na porcentagem
                if sla_percent >= 95:
                    color = '#22c55e'  # Verde (Excelente)
                elif sla_percent >= 90:
                    color = '#eab308'  # Amarelo (Bom)
                elif sla_percent >= 85:
                    color = '#f59e0b'  # Laranja (Regular)
                else:
                    color = '#ef4444'  # Vermelho (Insuficiente)
                
                result_data.append({
                    'ferramenta': tool_name,
                    'sla_percent': round(sla_percent, 1),
                    'color': color
                })
        
        # Ordena o resultado do melhor para o pior SLA
        result_data.sort(key=lambda x: x['sla_percent'], reverse=True)
        
        return jsonify(result_data)
        
    except Exception as e:
        logger.error(f"Erro ao gerar SLA real por ferramentas: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


      

@app.route('/export_comments', methods=['POST'])
@login_required
@permission_required
def export_comments():
    try:
        logger.debug("Requisição recebida para /export_comments")
        csr_list = request.form.get('csr_list')
        export_type = request.form.get('export_type')
        logger.debug(f"CSR List: {csr_list}, Export Type: {export_type}")
        if not csr_list or not export_type:
            logger.error("CSR ou export_type não fornecidos")
            return "Lista de CSRs e tipo de exportação são obrigatórios.", 400
        csrs = [csr.strip() for csr in csr_list.split(',') if csr.strip()]
        logger.debug(f"CSRs processadas: {csrs}")
        excel_file_path = r"E:/Python/Coleta-Site/exportado_formatado.xlsx"
        logger.debug(f"Carregando arquivo: {excel_file_path}")
        df = pd.read_excel(excel_file_path)
        logger.debug(f"Colunas no Excel: {list(df.columns)}")
        if 'Comentários' not in df.columns:
            logger.error("Coluna 'Comentários' não encontrada no arquivo Excel")
            return "Coluna 'Comentários' não encontrada no arquivo Excel", 400
        df_filtered = df[df['URL'].str.contains('|'.join(csrs), case=False, na=False)].copy()
        logger.debug(f"Linhas após filtro por CSRs: {len(df_filtered)}")
        if df_filtered.empty:
            logger.warning(f"Nenhum comentário encontrado para as CSRs: {csr_list}")
            return f"Nenhum comentário encontrado para as CSRs {csr_list}.", 404
        if 'Timestamp' in df.columns:
            df_filtered['Timestamp'] = pd.to_datetime(df_filtered['Timestamp'], errors='coerce').dt.tz_localize(None).dt.tz_localize(ZoneInfo("America/Sao_Paulo"))
            df_filtered = df_filtered.sort_values(by='Timestamp', ascending=True)
            logger.debug("Ordenado por Timestamp em America/Sao_Paulo")
        else:
            logger.debug("Coluna 'Timestamp' não encontrada, usando ordem do arquivo")
        if export_type == 'last':
            df_filtered['CSR'] = df_filtered['URL'].str.extract(r'(\d+)$')
            df_result = pd.DataFrame()
            for csr in csrs:
                df_csr = df_filtered[df_filtered['URL'].str.contains(csr, case=False, na=False)].tail(1)
                if not df_csr.empty:
                    last_row = df_csr.iloc[0]
                    comentarios = last_row['Comentários'].split('👤')
                    ultimo_comentario = comentarios[-1].strip() if len(comentarios) > 1 else last_row['Comentários']
                    df_result = pd.concat([df_result, pd.DataFrame([{
                        'URL': last_row['URL'],
                        'Comentários': f"👤 {ultimo_comentario}",
                        'Status': last_row['Status'],
                        'Timestamp': last_row.get('Timestamp', '') if 'Timestamp' in last_row else ''
                    }])], ignore_index=True)
            logger.debug("Exportando apenas o último comentário de cada CSR")
            df_filtered = df_result
        elif export_type == 'all':
            logger.debug("Exportando todos os comentários")
        else:
            logger.error(f"Tipo de exportação inválido: {export_type}")
            return "Tipo de exportação inválido.", 400
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_filtered.to_excel(writer, index=False, sheet_name='Comentários')
        output.seek(0)
        logger.debug("Arquivo Excel gerado com sucesso")
        filename = f"comentarios_csr_{'_'.join(csrs)}.xlsx"
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return f"Ocorreu um erro ao processar sua solicitação: {str(e)}", 500

def gerar_os():
    agora = datetime.now(ZoneInfo("America/Sao_Paulo"))
    ano = str(agora.year)[-2:]
    dia_do_ano = f"{agora.timetuple().tm_yday:03}"
    segundos = f"{agora.second:02}"
    numeros_aleatorios = f"{random.randint(0, 9999):04}"
    return f"{ano}{dia_do_ano}{segundos}{numeros_aleatorios}"

@app.route('/new_ticket', methods=['GET', 'POST'])
@login_required
@permission_required
def new_ticket():
    try:
        if request.method == 'POST':
            username = current_user.username
            tool = request.form['tool']
            subject = request.form['subject']
            description = request.form['description']
            csr = request.form.get('csr', None)
            priority = request.form['priority']

            os_numero = gerar_os()

            new_ticket = Ticket(
                username=username,
                tool=tool,
                subject=subject,
                description=description,
                csr=csr,
                priority=priority,
                os=os_numero
            )
            db.session.add(new_ticket)
            db.session.commit()

            files = request.files.getlist('files')
            for file in files:
                if file:
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    new_file = File(ticket_id=new_ticket.id, filename=filename)
                    db.session.add(new_file)

            db.session.commit()
            flash(f'Ticket criado com sucesso! OS: {os_numero}', 'success')
            return redirect(url_for('index'))

        return render_template('new_ticket.html', username=current_user.username)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/uploads/<filename>')
@login_required
@permission_required
def uploaded_file(filename):
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/ticket/<int:ticket_id>/reopen', methods=['POST'])
@login_required
@permission_required
def reopen_ticket(ticket_id):
    try:
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket não encontrado.', 'error')
            return redirect(url_for('index'))
        ticket.status = 'open'
        ticket.closed_at = None
        db.session.commit()
        flash('Ticket reaberto com sucesso.', 'success')
        return redirect(url_for('ticket_details', ticket_id=ticket.id))
    except Exception as e:
        logger.error(f"Erro: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Garante que o usuário admin exista
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', is_admin=True)
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            
    # Rodando no IP e Porta solicitados

    app.run(host='10.32.36.194', port=5000, debug=True)
